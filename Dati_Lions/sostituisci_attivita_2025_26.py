"""
Sostituzione MIRATA delle attività dell'anno sociale 2025-26 in attivita_report,
a partire dall'export Lions xlsx (filtrato per data creazione AF 2025-26).

Piano (solo con --go; senza --go = dry-run in sola lettura):
  1. BACKUP: create table bak_attreport_<ts> as select * from attivita_report
  2. UPSERT per id_attivita (aggiorna esistenti, inserisce nuove) di TUTTE le
     righe del file (2426). Le 21 con data_inizio 2024-25 finiscono nel loro anno.
  3. stato_approvazione = 'approvato' per tutti gli id importati (scelta utente).
  4. DELETE "stale": righe con data_inizio in [2025-07-01, 2026-06-30] il cui
     id NON è nel file (attività 2025-26 sparite dal nuovo export).
  5. Verifiche conteggi.

Token Management API: env SUPABASE_ACCESS_TOKEN oppure Dati_Lions/.mgmt_token.
"""
import json, os, sys, urllib.request, urllib.error, datetime
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__)); os.chdir(HERE)
GO = "--go" in sys.argv
PROJECT = "uywtfwjkyiacdfgsbtgo"
TOKEN = (os.environ.get("SUPABASE_ACCESS_TOKEN") or "").strip()
if not TOKEN and os.path.exists(".mgmt_token"):
    TOKEN = open(".mgmt_token", encoding="utf-8").read().strip()
if not TOKEN:
    raise SystemExit("ERRORE: manca il token (.mgmt_token o SUPABASE_ACCESS_TOKEN).")
URL = f"https://api.supabase.com/v1/projects/{PROJECT}/database/query"

XLSX = os.environ.get("XLSX_PATH") or r"C:\Users\f.coppini\Downloads\Service Activities Information-2026-07-22-11-39-50.xlsx"
Y_FROM, Y_TO = "2025-07-01", "2026-06-30"

def exec_sql(sql):
    body = json.dumps({"query": sql}).encode("utf-8")
    req = urllib.request.Request(URL, data=body, method="POST", headers={
        "Authorization": f"Bearer {TOKEN}", "Content-Type": "application/json", "User-Agent": "Mozilla/5.0"})
    try:
        with urllib.request.urlopen(req, timeout=120) as resp:
            return json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        raise RuntimeError(f"HTTP {e.code}: {e.read().decode('utf-8')[:600]}")

# ---- colonne tabella (come 4_report_supabase.csv) ----
HEADERS = ["id_attivita","id_account_club","nome_club_sponsor","titolo","descrizione","stato",
    "rapporto_completo","livello_attivita","causa","tipo_progetto","attivita_distintiva","finanziata_lcif",
    "data_inizio","data_conclusione","persone_servite","persone_servite_limite","totale_volontari",
    "totale_ore_servizio","totale_ore_servizio_capped","alberi_piantati","valuta_fondi_donati",
    "totale_fondi_donati","fondi_donati_usd_capped","donazione_lcif","organizzazione_beneficiata",
    "valuta_fondi_raccolti","totale_fondi_raccolti","fondi_raccolti_usd_capped","creato_da"]
BOOL_COLS = {"rapporto_completo","attivita_distintiva","finanziata_lcif"}
NUM_COLS = {"persone_servite","persone_servite_limite","totale_volontari","totale_ore_servizio",
    "totale_ore_servizio_capped","alberi_piantati","totale_fondi_donati","fondi_donati_usd_capped",
    "donazione_lcif","totale_fondi_raccolti","fondi_raccolti_usd_capped"}
DATE_COLS = {"data_inizio","data_conclusione"}
# indice colonna xlsx (header riga 22) -> campo tabella
COLMAP = {29:"id_attivita",33:"id_account_club",3:"nome_club_sponsor",9:"titolo",10:"descrizione",8:"stato",
    7:"rapporto_completo",11:"livello_attivita",12:"causa",13:"tipo_progetto",14:"attivita_distintiva",
    15:"finanziata_lcif",5:"data_inizio",6:"data_conclusione",16:"persone_servite",17:"persone_servite_limite",
    18:"totale_volontari",19:"totale_ore_servizio",20:"totale_ore_servizio_capped",27:"alberi_piantati",
    21:"totale_fondi_donati",22:"fondi_donati_usd_capped",23:"donazione_lcif",24:"organizzazione_beneficiata",
    25:"totale_fondi_raccolti",26:"fondi_raccolti_usd_capped",28:"creato_da"}

def norm_date(v):
    if v in (None,""): return None
    s = str(v).strip()
    if "/" in s:  # M/D/YYYY
        try:
            m,d,y = [int(x) for x in s.split("/")]; return f"{y:04d}-{m:02d}-{d:02d}"
        except Exception: return None
    if len(s) >= 10 and s[4]=="-" and s[7]=="-": return s[:10]
    return None

def anno_soc(iso):
    if not iso: return None
    y=int(iso[:4]); m=int(iso[5:7]); return y if m>=7 else y-1

def to_field(col, raw):
    if col in DATE_COLS: return norm_date(raw)
    if raw is None: return None
    if col in BOOL_COLS: return True if raw in (True,"True","true",1) else (False if raw in (False,"False","false",0) else None)
    if col == "donazione_lcif": return 1 if raw in (True,"True","true",1) else 0
    if col in NUM_COLS:
        try: return float(raw)
        except Exception: return None
    s = str(raw).strip()
    if col == "organizzazione_beneficiata" and s.lower() in ("not specified",""): return None
    return s if s != "" else None

def load_rows():
    wb = load_workbook(XLSX, read_only=True, data_only=True); ws = wb.active
    allrows = list(ws.iter_rows(values_only=True))
    hi = next(i for i in range(len(allrows)) if sum(1 for c in allrows[i] if c not in (None,"")) > 8)
    out = []
    for r in allrows[hi+1:]:
        if r[29] in (None,""):  # senza ID attività = riga vuota/totale
            continue
        rec = {h: None for h in HEADERS}
        for idx, col in COLMAP.items():
            rec[col] = to_field(col, r[idx])
        out.append(rec)
    return out

def sql_val(col, v):
    if v is None: return "NULL"
    if col in BOOL_COLS: return "TRUE" if v else "FALSE"
    if col in NUM_COLS: return repr(float(v))
    if col in DATE_COLS: return f"'{v}'"
    return "'" + str(v).replace("\\","\\\\").replace("'","''") + "'"

def build_upsert(chunk):
    cols = ", ".join(HEADERS)
    vals = ",\n".join("(" + ", ".join(sql_val(h, r[h]) for h in HEADERS) + ")" for r in chunk)
    upd = ", ".join(f"{c}=EXCLUDED.{c}" for c in HEADERS if c != "id_attivita")
    return f"INSERT INTO attivita_report ({cols}) VALUES\n{vals}\nON CONFLICT (id_attivita) DO UPDATE SET {upd};"

def q1(sql, key="n"):
    r = exec_sql(sql)
    return r[0][key] if isinstance(r, list) and r else r

def main():
    rows = load_rows()
    ids = [r["id_attivita"] for r in rows]
    n2526 = sum(1 for r in rows if anno_soc(r["data_inizio"]) == 2025)
    other = len(rows) - n2526
    print(f"File: {len(rows)} attività | data_inizio 2025-26: {n2526} | altri anni: {other}")
    print(f"ID unici: {len(set(ids))}")

    cur_tot = q1("SELECT COUNT(*)::int AS n FROM attivita_report;")
    cur_2526 = q1(f"SELECT COUNT(*)::int AS n FROM attivita_report WHERE data_inizio BETWEEN '{Y_FROM}' AND '{Y_TO}';")
    idlist = ",".join("'" + i.replace("'","''") + "'" for i in ids)
    gia = q1(f"SELECT COUNT(*)::int AS n FROM attivita_report WHERE id_attivita IN ({idlist});")
    stale = q1(f"SELECT COUNT(*)::int AS n FROM attivita_report WHERE data_inizio BETWEEN '{Y_FROM}' AND '{Y_TO}' AND id_attivita NOT IN ({idlist});")
    print(f"DB ora: totale={cur_tot} | 2025-26={cur_2526}")
    print(f"  del file, già in DB (update): {gia} | nuove (insert): {len(ids)-gia}")
    print(f"  2025-26 in DB che SPARIRANNO (stale delete): {stale}")

    if not GO:
        print("\n[DRY-RUN] nessuna scrittura. Rilancia con --go per eseguire.")
        return

    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M")
    bak = f"bak_attreport_{ts}"
    print(f"\n[1/5] BACKUP -> {bak}")
    exec_sql(f"CREATE TABLE {bak} AS SELECT * FROM attivita_report;")
    bn = q1(f"SELECT COUNT(*)::int AS n FROM {bak};")
    assert bn == cur_tot, f"backup incoerente: {bn} != {cur_tot}"
    print(f"      ok, {bn} righe salvate")

    print("[2/5] UPSERT per id_attivita")
    B = 120
    for i in range(0, len(rows), B):
        exec_sql(build_upsert(rows[i:i+B]))
        sys.stdout.write(f"      {min(i+B,len(rows))}/{len(rows)}\r"); sys.stdout.flush()
    print()

    print("[3/5] stato_approvazione = 'approvato' per gli id importati")
    for i in range(0, len(ids), 400):
        chunk = ",".join("'" + x.replace("'","''") + "'" for x in ids[i:i+400])
        exec_sql(f"UPDATE attivita_report SET stato_approvazione='approvato' WHERE id_attivita IN ({chunk});")

    print("[4/5] DELETE stale (2025-26 non presenti nel file)")
    d = exec_sql(f"WITH del AS (DELETE FROM attivita_report WHERE data_inizio BETWEEN '{Y_FROM}' AND '{Y_TO}' AND id_attivita NOT IN ({idlist}) RETURNING 1) SELECT COUNT(*)::int AS n FROM del;")
    print(f"      eliminate: {d}")

    print("[5/5] VERIFICHE")
    post_tot = q1("SELECT COUNT(*)::int AS n FROM attivita_report;")
    post_2526 = q1(f"SELECT COUNT(*)::int AS n FROM attivita_report WHERE data_inizio BETWEEN '{Y_FROM}' AND '{Y_TO}';")
    appr = q1(f"SELECT COUNT(*)::int AS n FROM attivita_report WHERE id_attivita IN ({idlist}) AND stato_approvazione='approvato';")
    print(f"      totale DB: {cur_tot} -> {post_tot}")
    print(f"      2025-26 in DB: {cur_2526} -> {post_2526} (atteso {n2526})")
    print(f"      importate approvate: {appr}/{len(ids)}")
    print(f"\nBackup: {bak} (per ripristino se serve). FATTO.")

if __name__ == "__main__":
    main()
